from __future__ import print_function
import ast
import xlrd  # reading Excel
from tsUserGroupApi import UsersAndGroups, User, Group, eprint
import json
from openpyxl import Workbook

"""
Copyright 2018 ThoughtSpot
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to
permit persons to whom the Software is furnished to do so, subject to the following conditions:
The above copyright notice and this permission notice shall be included in all copies or substantial portions
of the Software.
THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED
TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""

# -------------------------------------------------------------------------------------------------------------------

"""Classes to work with the TS public user and list APIs"""

# Helper functions. ----------------------------------------------------------------------


class UGXLSWriter(object):
    """
    Writes users and groups to an Excel spreadsheet.
    """

    def write(self, users_and_groups, filename):
        """
        Writes the content to the given file.
        :param users_and_groups:  The UsersAndGroups object to write.
        :type users_and_groups: UsersAndGroups
        :param filename:  Name of the file to write to.  No extension is expected and one will be added.
        :type filename: str
        """
        workbook = Workbook()
        workbook.remove_sheet(
            workbook.active
        )  # remove the default sheet since we'll be creating the ones we want.
        self._write_users(workbook, users_and_groups.get_users())
        self._write_groups(workbook, users_and_groups.get_groups())
        if not (filename.endswith("xls") or filename.endswith("xlsx")):
            filename += ".xlsx"

        workbook.save(filename)

    def _write_users(self, workbook, users):
        """
        Writes the users to a worksheet.
        :param workbook:  The workbook to write to.
        :type workbook:  Workbook
        :param users:  The list of groups to write.
        :type users: list of User
        :return:
        """
        ws = workbook.create_sheet(title="Users")
        self._write_header(
            ws,
            [
                "Name",
                "Password",
                "Display Name",
                "Email",
                "Groups",
                "Visibility",
                "Created",
                "ID"
            ],
        )
        cnt = 2  # start after header.
        for user in users:
            ws.cell(column=1, row=cnt, value=user.name)
            ws.cell(column=2, row=cnt, value=user.password)
            ws.cell(column=3, row=cnt, value=user.displayName)
            ws.cell(column=4, row=cnt, value=user.mail)
            ws.cell(column=5, row=cnt, value=json.dumps(user.groupNames))
            ws.cell(column=6, row=cnt, value=user.visibility)
            ws.cell(column=7, row=cnt, value=user.created)
            ws.cell(column=8, row=cnt, value=user.id)
            cnt += 1

    def _write_groups(self, workbook, groups):
        """
        Writes the groups to a worksheet.
        :param workbook:  The workbook to write to.
        :type workbook:  Workbook
        :param groups:  The list of groups to write.
        :type groups: list
        :return:
        """
        ws = workbook.create_sheet(title="Groups")
        self._write_header(
            ws,
            [
                "Name",
                "Display Name",
                "Description",
                "Groups",
                "Visibility",
                "Created",
            ],
        )
        cnt = 2  # start after header.
        for group in groups:
            ws.cell(column=1, row=cnt, value=group.name)
            ws.cell(column=2, row=cnt, value=group.displayName)
            ws.cell(column=3, row=cnt, value=group.description)
            ws.cell(column=4, row=cnt, value=json.dumps(group.groupNames))
            ws.cell(column=5, row=cnt, value=group.visibility)
            ws.cell(column=6, row=cnt, value=group.created)
            cnt += 1

    def _write_header(self, worksheet, cols):
        """
        Writes the header for the given worksheet in row 1.
        :param worksheet:  Worksheet to write to.
        :param cols:  List of columns to write.
        """
        for ccnt in range(0, len(cols)):
            worksheet.cell(column=(ccnt + 1), row=1, value=cols[ccnt])


class UGXLSReader(object):
    """
    Reads user and group info from an Excel file that is formatted the same as the UGXLSWriter writes.
    """

    required_sheets = ["Users", "Groups"]
    required_columns = {
        "Users": [
            "Name",
            "Password",
            "Display Name",
            "Email",
            "Groups",
            "Visibility"
        ],
        "Groups": [
            "Name",
            "Display Name",
            "Description",
            "Groups",
            "Visibility"
        ],
    }

    def __init__(self):
        """
        Creates a new UGXLSReader
        """
        self.workbook = None
        self.indices = {}
        self.users_and_groups = UsersAndGroups()

    def read_from_excel(self, filepath):
        """
        Reads users and groups from the given file.
        :param filepath:  Path to the Excel file to read from.
        :type filepath: str
        :return: Returns the users and groups read from the Excel file.  The users and groups are not validated
        :rtype UsersAndGroups
        so that they can be modified prior to validation.
        """
        self.workbook = xlrd.open_workbook(filepath)
        if self._verify_file_format():
            self._get_column_indices()
            self._read_users_from_workbook()
            self._read_groups_from_workbook()
        return self.users_and_groups

    def _verify_file_format(self):
        """
        :return: True if the format of the workbook is valid.
        :rtype: bool
        """
        is_valid = True
        sheet_names = self.workbook.sheet_names()
        for required_sheet in UGXLSReader.required_sheets:
            if required_sheet not in sheet_names:
                eprint("Error:  missing sheet %s!" % required_sheet)
                is_valid = False
            else:
                sheet = self.workbook.sheet_by_name(required_sheet)
                header_row = sheet.row_values(rowx=0, start_colx=0)
                for required_column in UGXLSReader.required_columns[
                    required_sheet
                ]:
                    if required_column not in header_row:
                        eprint(
                            "Error:  missing column %s in sheet %s!"
                            % (required_column, required_sheet)
                        )
                        is_valid = False

        return is_valid

    def _get_column_indices(self):
        """
        Reads the sheets to get all of the column indices.  Assumes the format was already checked.
        """
        sheet_names = self.workbook.sheet_names()
        for sheet_name in sheet_names:
            if sheet_name in self.required_sheets:
                sheet = self.workbook.sheet_by_name(sheet_name)
                col_indices = {}
                ccnt = 0
                for col in sheet.row_values(rowx=0, start_colx=0):
                    col_indices[col] = ccnt
                    ccnt += 1
                self.indices[sheet_name] = col_indices

    def _read_users_from_workbook(self):
        """
        Reads all the users from the workbook.
        """

        table_sheet = self.workbook.sheet_by_name("Users")
        indices = self.indices["Users"]

        for row_count in range(1, table_sheet.nrows):
            row = table_sheet.row_values(rowx=row_count, start_colx=0)

            # "Name", "Password", "Display Name", "Email", "Description", "Groups", "Visibility"
            username = row[indices["Name"]]
            password = row[indices["Password"]]
            display_name = row[indices["Display Name"]]
            email = row[indices["Email"]]
            groups = []
            if row[indices["Groups"]] and row[
                indices["Groups"]
            ]:
                groups = ast.literal_eval(
                    row[indices["Groups"]]
                )  # assumes a valid list format, e.g. ["a", "b", ...]
            visibility = row[indices["Visibility"]]

            try:
                user = User(
                    name=username,
                    password=password,
                    display_name=display_name,
                    mail=email,
                    group_names=groups,
                    visibility=visibility,
                )
                # The format should be consistent with only one user per line.
                self.users_and_groups.add_user(
                    user, duplicate=UsersAndGroups.RAISE_ERROR_ON_DUPLICATE
                )
            except:
                eprint("Error reading user with name %s" % username)

    def _read_groups_from_workbook(self):
        """
        Reads all the groups from the workbook.
        """

        table_sheet = self.workbook.sheet_by_name("Groups")
        indices = self.indices["Groups"]

        for row_count in range(1, table_sheet.nrows):
            row = table_sheet.row_values(rowx=row_count, start_colx=0)

            # Name", "Display Name", "Description", "Groups", "Visibility"
            group_name = row[indices["Name"]]
            display_name = row[indices["Display Name"]]
            description = row[indices["Description"]]
            visibility = row[indices["Visibility"]]

            groups = []
            if row[indices["Groups"]] and row[
                indices["Groups"]
            ]:
                groups = ast.literal_eval(
                    row[indices["Groups"]]
                )  # assumes a valid list format, e.g. ["a", "b", ...]
            try:
                group = Group(
                    name=group_name,
                    display_name=display_name,
                    description=description,
                    group_names=groups,
                    visibility=visibility,
                )
                # The format should be consistent with only one group per line.
                self.users_and_groups.add_group(
                    group, duplicate=UsersAndGroups.RAISE_ERROR_ON_DUPLICATE
                )
            except:
                eprint("Error reading group with name %s" % group_name)
