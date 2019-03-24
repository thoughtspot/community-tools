"""
Copyright 2017 ThoughtSpot

Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to
permit persons to whom the Software is furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all copies or substantial portions
of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED
TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""

import re
import sys
import xlrd  # reading Excel
import logging
import contextlib

# import datetime -- used by TsloadWriter, but commented out for now.
import csv
import os
from openpyxl import Workbook  # writing Excel
from datamodel import Database, Table, Column, ShardKey, DatamodelConstants, eprint
from tqlgenerator import TQLCommandGenerator, list_to_string

# -------------------------------------------------------------------------------------------------------------------


@contextlib.contextmanager
def smart_open(filename=None):
    """
    Borrowed from https://stackoverflow.com/questions/17602878/how-to-handle-both-with-open-and-sys-stdout-nicely
    :param filename: Name of the file to write to or '-' for stdout.
    """
    if filename and filename != '-':
        fh = open(filename, 'w')
    else:
        fh = sys.stdout

    try:
        yield fh
    finally:
        if fh is not sys.stdout:
            fh.close()


# -------------------------------------------------------------------------------------------------------------------


class DDLParser(object):
    """
    Parses DDL from various formats and creates a DataModel object that can be used for writing data.
    The following assumptions are made about the DDL being read:
    * Statements end with a semi-colon (;).
    * CREATE TABLE occur together on a single line, not split across lines.
    * CREATE TABLE statements will not occur inside of a comment block.
    * Delimiters, such as commas, will not be part of the table or column name.
    * Comment characters, such as #, --, or /* */ will not be part of a column name.
    * CREATE TABLE will have (....) with no embedded, unbalanced parentheses.
    """

    def __init__(
        self,
        database_name,
        schema_name=DatamodelConstants.DEFAULT_SCHEMA
    ):
        """
        Creates a new DDL parser.
        :param database_name: Name of the database to create.
        :type database_name: str
        :param schema_name: Name of the schema if not using the default.
        :type schema_name: str
        """
        self.schema_name = schema_name
        self.database_name = database_name
        self.database = None  # set when parsing.

    def parse_ddl(self, filename):
        """
        Parsed DDL from a stream and returns a populated Database.
        :param filename: Name of the file to read from.
        :return: A Database object.
        :rtype: Database
        """
        # Reset for new parse job.
        self.database = Database(self.database_name)

        # First read the entire input into memory.  This will allow multiple passes through the data.
        statements = self._get_statements(filename=filename)
        for stmt in statements:
            logging.debug(">>> %s" % stmt)
            lower = stmt.lower()
            if "create database" in lower:
                logging.debug("Ignoring create database statement.")
            elif "create table" in lower or "create or replace table" in lower:
                self._parse_create_table(stmt)
            elif "alter table" in lower:
                logging.debug("altering a table....")
                if "primary key" in lower:
                    logging.debug("adding a primary key:  %s" % stmt)
                    self._add_primary_key(stmt)
                elif "foreign" in lower:
                    logging.debug("adding foreign key:  %s" % stmt)
                    self._add_foreign_key(stmt)
                elif "relationship" in lower:
                    logging.debug("creating a relationship:  %s" % stmt)
                    self._add_generic_relationship(stmt)
                elif "hash" in lower:
                    logging.debug("sharding a table:  %s" % stmt)
                    self._add_shard_key(stmt)
                else:
                    logging.debug("ignoring alter")
            else:
                logging.debug("ignoring statement:  %s" % stmt)

        return self.database

    def _get_statements(self, filename):
        """
        Reads the statements from the file or input stream.
        :param filename:
        :return:
        """

        if filename is None:
            ddl_file = open(sys.stdin, "r")
        else:
            ddl_file = open(filename, "r")

        statements = []
        try:
            stmt_buffer = ""  # line stmt_buffer for reading entire commands.
            in_comment = False
            for line in ddl_file:
                line = self._clean_line(line=line)
                line = line.partition("--")[0]  # strip off any -- comments.
                logging.debug(line)
                if self._should_ignore_line(line=line):
                    pass
                else:
                    # if in a comment, then ignore lines until the comment is gone.
                    if in_comment:
                        # get content if done with comment.
                        if "*/" in line:
                            in_comment = False
                            line = line.partition("*/")[2]
                            stmt_buffer += line
                        # will ignore lines inside a comment.

                    elif "/*" in line:
                        # have an opening for comments.  Read until get to end of the comment.
                        in_comment = True
                        before_comment = line.partition("/*")[0]
                        after_comment = line.partition("/*")[2]
                        stmt_buffer += before_comment

                        if "*/" in after_comment:  # could be a one line comment.
                            in_comment = False
                            stmt_buffer += after_comment.partition("*/")[2]
                    else:
                        stmt_buffer += " " + line

                    if ";" in stmt_buffer:
                        # contains end of statement.
                        parts = stmt_buffer.partition(";")
                        statements.append(parts[0].strip())
                        logging.debug("adding statement:  %s" % parts[0])
                        stmt_buffer = parts[2]

        except Exception as ex:
            eprint(ex)
        finally:
            ddl_file.close()

        statements = [re.sub(" +", " ", stmt) for stmt in statements]

        return statements

    def _should_ignore_line(self, line):
        """
        Returns true if a line should be ignored.
        :param line: The line to test.
        :type line: str
        :return: True if should be ignored.
        """
        should_ignore = False

        # SQL Server cases.
        if line.startswith("GO"):
            should_ignore = True

        return should_ignore

    def _parse_create_table(self, statement):
        """
        Parses a create table statement.
        :param statement: The statement read in.
        :type statement: str
        :return:
        """
        statement = statement.replace("[", '"').replace(
            "]", '"'
        )  # for SQL Server quotes
        table_name = self._get_table_name(statement)
        table = Table(table_name=table_name, schema_name=self.schema_name)
        self._add_columns(table, statement)
        if "partition by hash" in statement.lower():
            self._add_hashkey(table, statement)

        self.database.add_table(table)

    def _get_table_name(self, statement):
        """
        Gets the table name from the statement.
        :param statement: The line with the create details.
        :type statement: str
        :return: The name of the table.
        :rtype: str
        """
        # The table name (and maybe a schema) are before the opening (
        tn = statement[0:statement.find("(")].rstrip()
        # strip off the first part of the statement to get just the name.
        strip_left_len = len("create table ")
        if "or replace" in tn.lower():
            strip_left_len = len("create or replace table ")
        tn = tn[strip_left_len:]
        # strip out schemas.
        tn = tn.split(".")[-1]
        tn = self._strip_quotes(tn)
        return tn

    @staticmethod
    def _strip_quotes(line):
        """
        Strips off any quotes in the given line.
        :param line: The line to strip quotes from.
        :type line: str
        :return: The line without quotes.
        :rtype: str
        """
        return line.replace("'", "").replace("`", "").replace('"', "")

    def _add_columns(self, table, statement):
        """
        Get the columns from the table statement.
        :param table: The table to add the columns to.
        :type table: Table
        :param statement: The statement with the create details.
        :type statement: str
        :return: A list of Columns
        :rtype: list
        """
        # The fields will be between the ( ).
        columns = []
        statement = statement[statement.find("(") + 1:statement.rfind(")")].strip()

        # think all DBs use commas for field separators
        # need to find the commas that are not inside of parents.
        field_statement = ""
        open_paren = False
        raw_fields = []

        for c in statement:

            if open_paren:
                field_statement += c
                if c == ")":
                    open_paren = False
            elif c == "(":
                field_statement += c
                open_paren = True
            else:
                if c == ",":
                    raw_fields.append(field_statement.strip())
                    field_statement = ""
                else:
                    field_statement += c

        if field_statement != "":
            raw_fields.append(field_statement.strip())

        for rf in raw_fields:  # get rid of any extraneous white space.
            rfl = rf.lower()

            # ignore key declarations.
            if "key " in rfl:
                if "primary" in rfl:
                    pks = rf.partition("(")[2].partition(")")[0].replace(" ", "").replace('"', '').split(",")
                    table.set_primary_key(pks)
                continue  # skip other key types and go to next field.

            had_quote = False
            if rfl[0] in "\"'`":  # should be a quote or letter
                had_quote = True
                name = rf[1:rf.find(rf[0], 1)]
            else:
                name = rf[0:rf.find(" ")]

            # The type comes after the name and goes up to the first of a
            #   space, close paren, or comma.  Assuming no space in type.
            start_idx = len(name) + (
                3 if had_quote else 1
            )  # extra 1 for space
            close_paren_idx = rfl.find(")", start_idx)
            if close_paren_idx > 0:  # type with ()
                data_type = rf[start_idx:close_paren_idx + 1]
            else:
                # either next space or comma.
                space_end_idx = rf.find(" ", start_idx)
                comma_end_idx = rf.find(",", start_idx)
                if space_end_idx == -1:  # not found
                    if comma_end_idx == -1:  # neither found
                        end_idx = len(rf)  # end of line
                    else:
                        end_idx = comma_end_idx
                elif comma_end_idx == -1:
                    end_idx = space_end_idx
                else:
                    end_idx = min(space_end_idx, comma_end_idx)
                data_type = rf[start_idx:end_idx]

            # print ("  adding %s as %s" % (name, data_type))
            columns.append(
                Column(
                    column_name=name, column_type=self._convert_type(data_type)
                )
            )

        table.add_columns(columns)

    @staticmethod
    def _convert_type(data_type):
        """
        Converts data types from other databases to ThoughtSpot types.
        :param data_type:  The datatype to convert.
        :type data_type: str
        :return: A ThoughtSpot data type.
        :rtype: str
        """
        if ")" in data_type:
            t = data_type[0:data_type.find(")") + 1]
        elif " " in data_type:
            t = data_type[0:data_type.find(" ") + 1]
        else:
            t = data_type

        t = t.lower()

        if "int" in t:
            new_t = "BIGINT"
        elif "rowversion" in t:  # MS type
            new_t = "INT"
        elif "uniqueidentifier" in t:  # Oracle type
            new_t = "VARCHAR(0)"
        elif "sysname" in t: # MS type
            new_t = "VARCHAR(0)"
        elif "serial" in t:  # serial index, Oracle and others
            new_t = "INT"
        elif "bit" in t:
            new_t = "BOOL"
        elif "blob" in t or "binary" in t:
            new_t = "UNKNOWN"
        elif "number" in t:  # support for NUMBER(1), NUMBER(1,1)
            if ")" in t:
                numsize = t[t.find("(") + 1:t.find(")")]
                if "," in numsize:
                    first_num, second_num = numsize.split(",")
                    if second_num.strip() == "0":
                        if first_num == "*" or int(
                            first_num
                        ) > 9:  # Support Oracle Number(*,n)
                            new_t = "BIGINT"
                        else:
                            new_t = "INT"
                    else:
                        new_t = "DOUBLE"
                else:
                    new_t = "INT"
            else:
                new_t = "BIGINT"
        elif "decimal" in t or "numeric" in t or "float" in t or "double" in t or "money" in t or "real" in t:
            new_t = "DOUBLE"
        elif "datetime" in t:
            new_t = "DATETIME"
        elif "timestamp" in t:
            new_t = "DATETIME"
        elif "time" in t:
            new_t = "TIME"
        elif "date" in t:
            new_t = "DATE"
        elif "bool" in t:
            new_t = "BOOL"
        elif "text" in t:
            new_t = "VARCHAR(0)"
        elif "long" in t:  # Oracle variable type
            new_t = "VARCHAR(0)"
        elif "enum" in t:
            new_t = "VARCHAR(0)"
        elif "xml" in t:
            new_t = "VARCHAR(0)"
        elif "char" in t:
            new_t = "VARCHAR(0)"
        # nbytes = 0
        # if ')' in t:
        #     nbytes = t[t.find('(') + 1:t.find(')')]
        #     nbytes = re.sub("[^0-9]", "", nbytes)
        #     if nbytes == "":
        #         nbytes = 0
        # new_t = "VARCHAR(%s)" % nbytes
        else:
            new_t = "UNKNOWN"

        return new_t

    @staticmethod
    def _clean_line(line):
        """
        Removes unwanted characters from the input line.
        :param line:  The line to clean up.
        :type line: str
        :return: The cleaned up line.
        :rtype: str
        """
        new_line = line.strip()
        new_line = re.sub(" +", " ", new_line)
        new_line = re.sub("\t+", " ", new_line)
        return new_line

    @staticmethod
    def _clean_name(name):
        """
        Strips out everything except names from a string.  If there are commas, those are left, but quotes, etc.
        are removed.  Also remove database or schema names.
        :param name: The name to clean up.
        :return: The cleaned name.
        """
        name = name.split(".")
        name = name[len(name)-1]
        cn = name.strip()
        cn = cn.replace("\t", "")
        cn = cn.replace("\"", "")
        cn = cn.replace("'", "")
        cn = cn.replace("[", "")
        cn = cn.replace("]", "")
        return cn

    @staticmethod
    def _add_hashkey(table, statement):
        """
        Reads the statement for a hashkey and adds to the table if it exists.  These will only be TQL hashkeys.
        :param table: The table to add the hashkey to, if it exists.
        :type table: Table
        :param statement: The create table statement.
        :type statement: str
        """
        try:
            pattern = re.compile("create table.*partition by hash.*\((.*)\) key.*\((.*)\)", re.IGNORECASE)
            # matches = pattern.search(statement.lower())
            matches = pattern.search(statement)
            number_shards = int(matches.group(1))
            shard_keys = matches.group(2)
            shard_keys = [DDLParser._clean_name(key) for key in shard_keys.split(",")]
            table.shard_key = ShardKey(shard_keys=shard_keys, number_shards=number_shards)
        except Exception as ex:
            eprint('Error "%s" extracting hash key from: %s' % (ex, statement))

    # -------------------------------------------------------------------------------------------------------------------
    @staticmethod
    def _extract_table_name(potential_name):
        """
        Extracts a table name.  It's possible that the name has a schema and other characters.
        :param potential_name: The value that has the table name.
        :return:  The table name.
        """
        # check for database, schema, etc.  Assume table is last in a dot separated list.
        table_name = potential_name.split(".")[-1]
        table_name = DDLParser._extract_name(table_name)
        return DDLParser._clean_name(table_name)

    @staticmethod
    def _extract_name(start_of_name):
        """
        Extracts a name from some text.  It assumes the name is at the start of the string and may or may not
        be enclosed in quotes.  If not in quotes, then the first token to the string is selected.  If in quotes, then the
        text between quotes is selected.
        :param start_of_name: String that has the start of the name.  This can be a quote or not.
        :type start_of_name: str
        :return: The name of the extraced from the text.
        """
        if start_of_name[0] == '"':
            return start_of_name[1:].split('"')[0]

        if start_of_name[0] == "'":
            return start_of_name[1:].split("'")[0]

        return start_of_name.split(" ")[0]

    @staticmethod
    def _get_matches(statement, patterns, expect_matches=1):
        """
        Returns a list of the matches in a statement based on a set of patterns.  The patterns will be tested one
        at a time and the first one that returns the expected number of matches is returned.
        :param statement: The statement to check for matches.
        :type statement: str
        :param patterns: List of patterns to use for matching.  Should be in order of preference since the first
        match will be returned.
        :type patterns: list of str
        :param expect_matches: number of matches being looked for.  Only matches with those many will be returned.
        :return: Either a tuple of matches or None if no appropriate match is found.
        :rtype: tuple
        """
        for p in patterns:
            match = re.match(p, statement, re.IGNORECASE)
            if match:
                if len(match.groups()) == expect_matches:
                    return match.groups()

        return None

    def _add_primary_key(self, statement):
        """
        Adds primary keys that are created with an alter table statement.
        primary keys can come in multiple patterns:
          TQL:  alter table .... primary key (keys...)
          SQL Server:  alter table .... primary key ... (keys...)
        :param statement: The alter table statement.
        :type statement: str
        """
        logging.debug("Trying to extract primary key from %s" % statement)
        try:
            patterns = ["alter table (.*) .*primary key.*\((.*)\)"]
            matches = DDLParser._get_matches(patterns=patterns, statement=statement, expect_matches=2)
            if not matches:
                eprint("Possible parsing error: unable to extract primary key from %s." % statement)
                return

            table_name = DDLParser._extract_table_name(matches[0])
            primary_key = matches[1]
            primary_keys = [DDLParser._clean_name(key) for key in primary_key.split(",")]

            table = self.database.get_table(table_name=table_name)
            if table:
                table.set_primary_key(primary_key=primary_keys)
            else:
                logging.error("Attempting to add a primary key to table %s, which is not in the database." % table_name)
        except Exception as ex:
            eprint('Error "%s" extracting primary key from: %s' % (ex, statement) )

    def _add_foreign_key(self, statement):
        """
        Adds foreign keys that are created with an alter table statement.
        primary keys can come in multiple patterns:
          with name:  alter table <table-name> add constraint <constraint-name> foreign key (<key-name>) references <table-name> (<key-name>);
          without name:  alter table <table-name> add constraint foreign key (<key-name>) references <table-name> (<key-name>);
        :param statement: The alter table statement.
        :type statement: str
        """
        try:
            # There can be multiple foreign key constraints in a single add constraint command, so need to split
            # and test each sub-pattern.
            constraints = re.split("constraint", statement, flags=re.IGNORECASE)
            patterns = ["alter table (.*) add.*"]
            matches = DDLParser._get_matches(patterns=patterns, statement=statement, expect_matches=1)
            table_name = DDLParser._clean_name(matches[0])
            for idx in range(1,len(constraints)):
                # gets rid of unnecessary commas and semi-colons if they exist.
                constraint = constraints[idx].strip().strip(",").strip(";").strip()

                # There can either be a name or not.  If there is, the matches change.
                patterns = ["(.*) foreign key *\((.*)\) references *(.*) \((.*)\)"]
                matches = DDLParser._get_matches(patterns=patterns, statement=constraint, expect_matches=4)
                if matches:
                    constraint_name = DDLParser._clean_name(matches[0])
                    from_columns = [DDLParser._clean_name(k) for k in matches[1].split(",")]
                    to_table = DDLParser._extract_table_name(matches[2])
                    to_columns = [DDLParser._clean_name(k) for k in matches[3].split(",")]
                else:
                    patterns = ["foreign key *\((.*)\) references *(.*) \((.*)\)"]
                    matches = DDLParser._get_matches(patterns=patterns, statement=constraint, expect_matches=3)
                    if not matches:
                        eprint("Possible parsing error: unable to extract foreign key from %s." % statement)
                        return
                    else:
                        constraint_name = None
                        # from_columns = DDLParser._clean_name(matches[0])
                        from_columns = [DDLParser._clean_name(k) for k in matches[0].split(",")]
                        to_table = DDLParser._extract_table_name(matches[1])
                        # to_columns = DDLParser._clean_name(matches[2])
                        to_columns = [DDLParser._clean_name(k) for k in matches[2].split(",")]

                table = self.database.get_table(table_name=table_name)
                if table:
                    table.add_foreign_key(from_keys=from_columns, to_table=to_table, to_keys=to_columns, name=constraint_name)
                else:
                    logging.error("Attempting to add a foreign key to table %s, which is not in the database." % table_name)
        except Exception as ex:
            eprint('Error "%s" extracting foreign key from: %s' % (ex, statement) )

    def _add_generic_relationship(self, statement):
        """
        Adds generic relationships that are created with an ALTER TABLE statement.  Should only be TQL.
        Format for the statement is:
            with name:  alter table <table-name add relationship <relationship-name> with <table-name> as <condition>
            without name:  alter table <table-name add relationship with <table-name> as <condition>
        :param statement: The alter table statement.
        :type statement: str
        """
        try:
            # There can either be a name or not.  If there is, the matches change.
            patterns = ["alter table (.*) add relationship (.*) with (.*) as (.*)"]
            matches = DDLParser._get_matches(patterns=patterns, statement=statement, expect_matches=4)
            if matches:
                table_name = DDLParser._extract_table_name(matches[0])
                constraint_name = DDLParser._clean_name(matches[1])
                to_table = DDLParser._extract_table_name(matches[2])
                conditions = matches[3]
            else:
                patterns = ["alter table (.*) add relationship with (.*) as (.*)"]
                matches = DDLParser._get_matches(patterns=patterns, statement=statement, expect_matches=3)
                if not matches:
                    eprint("Possible parsing error: unable to extract generic relationship from %s." % statement)
                    return
                else:
                    table_name = DDLParser._extract_table_name(matches[0])
                    constraint_name = None
                    to_table = DDLParser._extract_table_name(matches[1])
                    conditions = matches[2]

            table = self.database.get_table(table_name=table_name)
            if table:
                table.add_relationship(to_table=to_table, name=constraint_name, conditions=conditions)
            else:
                logging.error("Attempting to add a relationship to table %s, which is not in the database." % table_name)
        except Exception as ex:
            eprint('Error "%s" extracting relationship from: %s' % (ex, statement) )

    def _add_shard_key(self, statement):
        """
        TODO implement this for alter statements.
        Adds shard keys (set partition).  TQL only.
        :param statement: The alter table statement.
        :type statement: str
        """
        try:
            # table_name = DDLParser._extract_name(matches.group(1))
            table_name = ""
            table = self.database.get_table(table_name=table_name)
            if table:
                pass
            else:
                logging.error("Attempting to add a shard key to table %s, which is not in the database." % table_name)
        except Exception as ex:
            eprint('Error "%s" extracting shard key from: %s' % (ex, statement) )


class TQLWriter:
    """
    Writes TQL from a data model.
    """

    def __init__(
        self,
        uppercase=False,
        lowercase=False,
        camelcase=False,
        create_db=False,
    ):
        """
        Creates a new TQLWriter to write database models to TQL.
        :param uppercase: Converts all names to uppercase.
        :type uppercase: bool
        :param lowercase: Converts all names to lowercase.  overrides uppercase if set.
        :type lowercase: bool
        :param create_db: Writes create statements for the database.
        :type create_db: bool
        """
        self.command_generator = TQLCommandGenerator(uppercase=uppercase, lowercase=lowercase, camelcase=camelcase)
        self.create_db = create_db

    def write_tql(self, database, filename=None):
        """
        Main function to write the Database to TQL.
        :param database: The database object to convert.
        :type database: Database
        :param filename: File to write to or STDOUT is not set.  The caller is expected to close the output stream.
        :type filename: str
        """

        with smart_open(filename) as outfile:

            db_name = database.database_name

            if self.create_db:
                outfile.write(self.command_generator.generate_create_database_statement(database_name=db_name))

            # sets to use the database.
            outfile.write(self.command_generator.generate_use_database_statement(database_name=db_name))

            # create the database if that was an option.
            if self.create_db:
                for schema_name in database.get_schema_names():
                    if schema_name != DatamodelConstants.DEFAULT_SCHEMA:
                        outfile.write(self.command_generator.generate_create_schema_statement(schema_name=schema_name))

            for table in database:
                self.write_create_table_statement(table, outfile)

            for table in database:
                self.write_foreign_keys(table, outfile)

            for table in database:
                self.write_relationships(table, outfile)

    def write_create_table_statement(self, table, outfile):
        """
        Writes a CREATE TABLE statement.
        :param table:  The table to create.
        :type table:  Table
        :param outfile: File stream to write to.
        """

        outfile.write("\n")
        outfile.write(self.command_generator.generate_drop_table_statement(table))
        outfile.write("\n")
        outfile.write(self.command_generator.generate_create_table_statement(table))

    def write_foreign_keys(self, table, outfile):
        """
        Writes alter table statements for foreign keys.  These should come after table creation.
        :param table: The table with the FK relationships.
        :type table: Table
        :param outfile: The file to write to.
        """
        for fk in table.foreign_keys.values():
            outfile.write(self.command_generator.generate_foreign_key_statement(table=table, foreign_key=fk))

    def write_relationships(self, table, outfile):
        """
        Writes alter table statements for relationships.  These should come after table creation.
        :param table: The table with the generic relationships.
        :type table: Table
        :param outfile: The file to write to.
        """
        for rel in table.relationships.values():
            outfile.write(self.command_generator.generate_relationships(table=table, relationship=rel))

# -------------------------------------------------------------------------------------------------------------------


class XLSWriter:
    """
    Writes data from a database to Excel.
    """
    EXTENSION = ".xlsx"

    def __init__(self):
        """
        Creates a new writer can write models to an Excel workbook.
        """
        self.workbook = Workbook()

    def write_database(self, database, filename):
        """
        Write the database to an Excel file.
        :param database:  The database object to write to Excel.
        :type database: Database
        :param filename:  Name of the Excel file without extension.
        :type filename: str
        """
        self._write_columns_worksheet(database)
        self._write_tables_worksheet(database)
        self._write_foreign_keys_worksheet(database)
        self._write_relationships_worksheet(database)
        self._write_to_excel(filename)

    def _write_columns_worksheet(self, database):
        """
        Writes the worksheet with the columns for each table.
        :param database: The database to write.
        :type database: Database
        """

        ws = self.workbook.active  # assumes this is the first sheet.
        ws.title = "Columns"

        # Write the header row.
        self._write_header(
            ws=ws,
            cols=["Database", "Schema", "Table", "Column", "Name", "Type"],
        )

        # Write the data.
        row_cnt = 1
        for table in database:
            col_idx = 0
            for column in table:
                col_idx += 1
                row_cnt += 1
                self._write_row(
                    ws=ws,
                    row_cnt=row_cnt,
                    cols=[
                        database.database_name,
                        table.schema_name,
                        table.table_name,
                        col_idx,
                        column.column_name,
                        column.column_type,
                    ],
                )

    def _write_tables_worksheet(self, database):
        """
        Writes the table(s) details to the table worksheet.
        :param database:  Database with the tables.
        :type database: Database
        """

        ws = self.workbook.create_sheet(title="Tables")
        self._write_header(
            ws=ws,
            cols=[
                "Database",
                "Schema",
                "Table",
                "Updated",
                "Update Type",
                "# Rows",
                "# Columns",
                "Primary Key",
                "Shard Key",
                "# Shards",
                "RLS Column",
                "# FKs From",
                "# FKs To"
            ],
        )
        # Write the data.
        row_cnt = 1
        for table in database:
            row_cnt += 1
            lookup_formula = "=COUNTIFS(Columns!A:A,Tables!A%d, Columns!B:B,Tables!B%d, Columns!C:C,Tables!C%d)" % (
                row_cnt, row_cnt, row_cnt
            )

            primary_key = ""
            if table.primary_key is not None:
                primary_key = list_to_string(table.primary_key)

            shard_key = ""
            number_shards = ""
            if table.shard_key is not None:
                shard_key = list_to_string(table.shard_key.shard_keys)
                number_shards = table.shard_key.number_shards

            # Formulas for seeing how many FKs are to and from the given table.
            nbr_fks_from = \
                "=IF(COUNTIF('Foreign Keys'!$D:$D,\"=\"&$C%d)>0,COUNTIF('Foreign Keys'!$D:$D,\"=\"&$C%d),\"\")" \
                % (row_cnt, row_cnt)
            nbr_fks_to = \
                "=IF(COUNTIF('Foreign Keys'!$F:$F,\"=\"&$C%d)>0,COUNTIF('Foreign Keys'!$F:$F,\"=\"&$C%d),\"\")" \
                % (row_cnt, row_cnt)

            # TODO add support for update frequency so that it's remembered during development.
            self._write_row(
                ws,
                row_cnt,
                [
                    database.database_name,
                    table.schema_name,
                    table.table_name,
                    "daily",
                    "partial",
                    "",
                    lookup_formula,
                    primary_key,
                    shard_key,
                    number_shards,
                    "",
                    nbr_fks_from,
                    nbr_fks_to
                ],
            )

    def _write_foreign_keys_worksheet(self, database):
        """
        Writes the relationship details (FKs and generic relationships) worksheet.
        :param database: 
        :type database: Database
        """
        ws = self.workbook.create_sheet(title="Foreign Keys")
        # Assuming relationships can only be within a single schema.
        self._write_header(
            ws,
            [
                "Name",
                "Database",
                "Schema",
                "From Table",
                "From Columns",
                "To Table",
                "To Columns",
            ],
        )

        row_cnt = 1
        for table in database:
            for fk in table.foreign_keys_iter():
                row_cnt += 1
                from_column = list_to_string(fk.from_keys)
                to_column = list_to_string(fk.to_keys)
                self._write_row(
                    ws,
                    row_cnt,
                    [
                        fk.name,
                        database.database_name,
                        table.schema_name,
                        fk.from_table,
                        from_column,
                        fk.to_table,
                        to_column,
                    ],
                )

    def _write_relationships_worksheet(self, database):
        """
        Writes the relationship details (FKs and generic relationships) worksheet.
        :param database: 
        :type database: Database
        """
        ws = self.workbook.create_sheet(title="Relationships")
        # Assuming relationships can only be within a single schema.
        self._write_header(
            ws,
            [
                "Name",
                "Database",
                "Schema",
                "From Table",
                "To Table",
                "Conditions",
            ],
        )

        row_cnt = 1
        for table in database:
            row_cnt += 1

            for rel in table.relationships_iter():
                self._write_row(
                    ws,
                    row_cnt,
                    [
                        rel.name,
                        database.database_name,
                        table.schema_name,
                        rel.from_table,
                        rel.to_table,
                        rel.conditions,
                    ],
                )

    @staticmethod
    def _write_header(ws, cols):
        """
        Writes the header for a worksheet.
        :param ws: The worksheet to write to.
        :param cols: The column headers to write.
        """
        for ccnt in range(0, len(cols)):
            ws.cell(column=(ccnt + 1), row=1, value=cols[ccnt])

    # TODO add formatting to distinguish the header.

    @staticmethod
    def _write_row(ws, row_cnt, cols):
        """
        Writes a row of data.
        :param ws: The worksheet to write to.
        :param row_cnt: The row to write to.
        :param cols: The columns to write.
        """
        for ccnt in range(0, len(cols)):
            ws.cell(column=(ccnt + 1), row=row_cnt, value=cols[ccnt])

    def _write_to_excel(self, filename):
        """
        Writes the data to an Excel file.
        :param filename: The name of the file (without extension) to write to.
        """
        if not filename.endswith(XLSWriter.EXTENSION):
            filename = filename + XLSWriter.EXTENSION

        self.workbook.save(filename)


# -------------------------------------------------------------------------------------------------------------------


class XLSReader:
    """
    Reads data models from an Excel file.  Note that this file follows a very specific format.  
    See test_excel_reader.xlsx for an example of the format.
    Note that this can return multiple databases. 
    Note that multiple tables with the same name in different schemas are not currently supported.
    """
    required_sheets = ["Columns", "Tables", "Foreign Keys", "Relationships"]
    required_columns = {
        "Columns": ["Database", "Schema", "Table", "Column", "Name", "Type"],
        "Tables": [
            "Database",
            "Schema",
            "Table",
            "Updated",
            "Update Type",
            "# Rows",
            "# Columns",
            "Primary Key",
            "Shard Key",
            "# Shards",
            "RLS Column",
        ],
        "Foreign Keys": [
            "Name",
            "Database",
            "Schema",
            "From Table",
            "From Columns",
            "To Table",
            "To Columns",
        ],
        "Relationships": [
            "Name",
            "Database",
            "Schema",
            "From Table",
            "To Table",
            "Conditions",
        ],
    }

    def __init__(self):
        """Creates a new Excel reader."""
        # Column indices for each sheet to make reading of values based on column header.
        self.workbook = None
        self.indices = {}
        self.databases = {}

    def read_xls(self, filepath):
        """
        Reads an Excel file at the given path and returns a dictionary with one or more database keyed by name.
        :param filepath: The path to the Excel document.
        :type filepath: str
        :return: A Database object based on the contents of the Excel document.
        :rtype: dict of str:Database
        """
        self.workbook = xlrd.open_workbook(filepath)
        if self._verify_file_format():
            self._read_databases_from_workbook()
        return self.databases

    def _verify_file_format(self):
        """
        Verifies that the Excel document has the correct tabs and column headers.
        :return: True if the workbook has the correct worksheets and columns in each.
        :rtype: bool
        """

        is_valid = True  # hope for the best.

        sheet_names = self.workbook.sheet_names()
        for required_sheet in XLSReader.required_sheets:
            if required_sheet not in sheet_names:
                eprint("Error:  missing sheet %s!" % required_sheet)
                is_valid = False
            else:
                sheet = self.workbook.sheet_by_name(required_sheet)
                header_row = sheet.row_values(rowx=0, start_colx=0)
                for required_column in XLSReader.required_columns[
                    required_sheet
                ]:
                    if required_column not in header_row:
                        eprint(
                            "Error:  missing column %s in sheet %s!"
                            % (required_column, required_sheet)
                        )
                        is_valid = False

        return is_valid

    def _get_column_indices(self):
        """
        Reads the sheets to get all of the column indices.  Assumes the format is valid.
        """

        sheet_names = self.workbook.sheet_names()
        for sheet_name in sheet_names:
            if sheet_name in self.required_sheets:
                sheet = self.workbook.sheet_by_name(sheet_name)
                col_indices = {}
                ccnt = 0
                for col in sheet.row_values(rowx=0, start_colx=0):
                    col_indices[col] = ccnt
                    ccnt += 1
                self.indices[sheet_name] = col_indices

    def _read_databases_from_workbook(self):
        """
        Reads one or more database from a workbook.  Errors can still be encountered, but the format should be OK.
        :return: A list of the databases in the file.
        :rtype: dict of Database
        """
        self._get_column_indices()
        self._read_tables_from_workbook()
        self._read_columns_from_workbook()
        self._read_foreign_keys_from_workbook()
        self._read_relationships_from_workbook()

    def _read_tables_from_workbook(self):
        """
        Reads the databases and tables from Excel.  These are used to populate from the remaining sheets.
        """

        # "Tables":        ["Database", "Schema", "Table", "Updated", "Update Type", "# Rows", "# Columns",
        #                   "Primary Key", "Shard Key", "# Shards", "RLS Column"],
        table_sheet = self.workbook.sheet_by_name("Tables")
        indices = self.indices["Tables"]

        for row_count in range(1, table_sheet.nrows):
            row = table_sheet.row_values(rowx=row_count, start_colx=0)

            database_name = row[indices["Database"]]
            if database_name == "": # ignore rows with no DBs.
                # eprint("Warning:  no database provided. Ignoring :" % row)
                continue

            database = self.databases.get(database_name, None)
            if database is None:
                database = Database(database_name=database_name)
                self.databases[database_name] = database

            pk = row[indices["Primary Key"]].strip()
            if pk == "":
                pk = None
            else:
                pk = [x.strip() for x in pk.split(",")]

            sk_name = row[indices["Shard Key"]].strip()
            sk_nbr_shards = row[indices["# Shards"]]

            if (sk_name == "" and sk_nbr_shards != "") or (
                sk_name != "" and sk_nbr_shards == ""
            ):
                eprint(
                    "ERROR:  %s need to provide both a shard key name and number of shards."
                    % row[indices["Table"]]
                )

            if sk_name == "":
                sk = None
            else:
                sk = [x.strip() for x in sk_name.split(",")]

            shard_key = None
            if sk_name and sk_nbr_shards:
                shard_key = ShardKey(
                    shard_keys=sk, number_shards=sk_nbr_shards
                )

            table = Table(
                table_name=row[indices["Table"]],
                schema_name=row[indices["Schema"]],
                primary_key=pk,
                shard_key=shard_key
            )
            database.add_table(table)

    def _read_columns_from_workbook(self):
        """
        Reads the columns for the tables from Excel.  
        """
        # "Columns":       ["Database", "Schema", "Table", "Column", "Name", "Type"],
        column_sheet = self.workbook.sheet_by_name("Columns")
        indices = self.indices["Columns"]

        for row_count in range(1, column_sheet.nrows):
            row = column_sheet.row_values(rowx=row_count, start_colx=0)

            database_name = row[indices["Database"]]
            database = self.databases.get(database_name, None)
            if database is None:
                eprint(
                    "ERROR:  Database %s from the Columns tab is not known."
                    % database_name
                )

            else:
                table_name = row[indices["Table"]]
                table = database.get_table(table_name)
                if table is None:
                    eprint(
                        "ERROR:  Table %s from the Columns tab is not known."
                        % table_name
                    )
                else:
                    table.add_column(
                        Column(
                            column_name=row[indices["Name"]],
                            column_type=row[indices["Type"]],
                        )
                    )

    def _read_foreign_keys_from_workbook(self):
        """
        Reads the foreign keys for the tables from Excel.  
        """

        # "Foreign Keys":  ["Name", "Database", "Schema", "From Table", "Columns", "To Table", "Columns"],
        fk_sheet = self.workbook.sheet_by_name("Foreign Keys")
        indices = self.indices["Foreign Keys"]

        for row_count in range(1, fk_sheet.nrows):
            row = fk_sheet.row_values(rowx=row_count, start_colx=0)

            database_name = row[indices["Database"]]
            database = self.databases.get(database_name, None)
            if database is None:
                eprint(
                    "ERROR:  Database %s from the Foreign Keys tab is not known."
                    % database_name
                )

            else:
                table_name = row[indices["From Table"]]
                table = database.get_table(table_name)
                if table is None:
                    eprint(
                        "ERROR:  Table %s from the Foreign Keys tab is not known."
                        % table_name
                    )

                else:
                    key_name = row[indices["Name"]]
                    if key_name is None:
                        eprint(
                            "ERROR:  Table %s from the Foreign Keys tab is missing a FK name."
                            % table_name
                        )

                    from_keys = row[indices["From Columns"]]
                    from_keys = [x.strip() for x in from_keys.split(",")]
                    to_keys = row[indices["To Columns"]]
                    to_keys = [x.strip() for x in to_keys.split(",")]
                    table.add_foreign_key(
                        name=key_name,
                        from_keys=from_keys,
                        to_table=row[indices["To Table"]],
                        to_keys=to_keys,
                    )

    def _read_relationships_from_workbook(self):
        """
        Reads the foreign keys for the tables from Excel.  
        """
        # "Relationships": ["Name", "Database", "Schema", "From Table", "To Table", "Conditions"]
        rel_sheet = self.workbook.sheet_by_name("Relationships")
        indices = self.indices["Relationships"]

        for row_count in range(1, rel_sheet.nrows):
            row = rel_sheet.row_values(rowx=row_count, start_colx=0)

            database_name = row[indices["Database"]]
            database = self.databases.get(database_name, None)
            if database is None:
                eprint(
                    "ERROR:  Database %s from the Relationships tab is not known."
                    % database_name
                )

            else:
                table_name = row[indices["From Table"]]
                table = database.get_table(table_name)
                if table is None:
                    eprint(
                        "ERROR:  Table %s from the Relationships tab is not known."
                        % table_name
                    )
                table.add_relationship(
                    to_table=row[indices["To Table"]],
                    conditions=row[indices["Conditions"]],
                )


# -------------------------------------------------------------------------------------------------------------------


class TsloadWriter:
    """
    Write the tsload commands with all the flags.
    NOTE:  This class is not complete.
    """

    def __init__(self, default_flags=None):
        """
        Create a tsload writer.
        :param default_flags: this includes the default flags used in tsload command.
        :type default_flags: dict
        """
        self._default_flags = {
            "empty_target": "",
            "max_ignored_rows": "0",
            "skip_second_fraction": "",
            "source_data_format": "",
            "null_value": "",
            "has_header_row": "",
        }
        if default_flags is not None:
            self._default_flags.update(default_flags)

    def write_tsloadcommand(self, database, filename):
        """
        Write tsload command to the file.
        :param database:  The database object to write to Excel.
        :type database: Database
        :param filename:  Name of the Excel file without extension.
        :type filename: str
        """
        with open(filename, "w") as tsload_outfile:

            for table in database:
                flags = self._get_flags_from_csv(database.database_name, table)
                tsload_string = "tsload "
                for flag, value in flags.items():
                    # TODO Add locic or flags without values.
                    tsload_string += '--%s "%s" ' % (flag, value)
                tsload_outfile.write(tsload_string + "\n")

    def _get_flags_from_csv(self, database_name, table):
        """
        This function will read the csv file and determine the flags
        :param database_name: Name of the database
        :type  database_name: str
        :param table: table object
        :type table: Table
        :return: all the dict for the tsload flags.
        :rtype: dict
        """
        flags = self._default_flags.copy()

        flags["target_database"] = database_name
        flags["target_schema"] = table.schema_name
        flags["target_table"] = table.table_name
        flags["source_file"] = table.table_name + ".csv"

        if not os.path.isfile(flags["source_file"]):
            return flags

        # TODO get default flags from csv

        with open(flags["source_file"]) as csv_file:
            csv_reader = csv.DictReader(csv_file)

            for column in table:
                if column.column_type == "DATE" and flags.get(
                    "date_format", None
                ) is None:
                    flags["date_format"] = TsloadWriter._get_date_format(
                        column.column_name, csv_reader
                    )

        return flags

    @staticmethod
    def _get_date_format(column_name, csv_reader):
        """
        :param column_name: Name of the column to get the date format for.
        :type column_name: str
        :param csv_reader: Open file reader.
        :type csv_reader: csv.DictReader
        :return: date_format for tsload command.
        :rtype: str
        """
        # TODO Add logic to determine the actual flag value.

        patterns = [
            "%Y",
            "%b %d, %Y",
            "%b %d, %Y",
            "%B %d, %Y",
            "%m/%d/%Y",
            "%Y/%m/%d",
            "%m/%d/%y",
            "%m-%d-%y",
            "%m-%d-%Y",
            "%Y-%m-%d",
            "%y/%m/%d",
            "%y-%m-%d",
            "%d/%m/%y",
            "%d/%m/%Y",
            "%d-%m-%y",
            "%d-%m-%Y",
            "%d%B%y",
            "%d%b%y",
            "%d%B%Y",
            "%d%b%Y",
        ]  # all the formats the date

        date_format = "%y/%m/%d"

        # parse = []

        # for fmt in patterns:
        #     try:
        #         dt = datetime.datetime.strptime(v, fmt)
        #         parse.append(fmt)
        #         print parse
        #         break

        #     except ValueError as err:
        #         pass

        # date_format = parse[0]

        return date_format
