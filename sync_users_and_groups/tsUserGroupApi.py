from __future__ import print_function
import sys
import ast
import requests
import logging
import json
import time
from collections import OrderedDict, namedtuple
from openpyxl import Workbook
import xlrd  # reading Excel
import requests.packages.urllib3
from requests.packages.urllib3.exceptions import InsecureRequestWarning

"""
Copyright 2018 ThoughtSpot
Permission is hereby granted, free of charge, to any person obtaining a copy of this software and associated
documentation files (the "Software"), to deal in the Software without restriction, including without limitation the
rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the Software, and to
permit persons to whom the Software is furnished to do so, subject to the following conditions:
The above copyright notice and this permission notice shall be included in all copies or substantial portions
of the Software.
THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED
TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
"""

# -------------------------------------------------------------------------------------------------------------------

"""Classes to work with the TS public user and list APIs"""

# Helper functions. ----------------------------------------------------------------------


def eprint(*args, **kwargs):
    """
    Prints to standard error similar to regular print.
    :param args:  Positional arguments.
    :param kwargs:  Keyword arguments.
    """
    print(*args, file=sys.stderr, **kwargs)


def public_props(obj):
    """
    Returns any property that doesn't start with an _
    """
    return (name for name in vars(obj).keys() if not name.startswith("_"))


def obj_to_json(obj):
    """
    Returns a json string with all of the objects public properties as attributes
    This function only goes one level deep and does not convert contents of lists, 
    dict, etc.
    :returns: A JSONS string representation of the object.
    :rtype: str
    """

    json_str = "{ "

    first = True
    for name in public_props(obj):
        value = getattr(obj, name)  # don't print empty values
        if not value:
            continue

        if first:
            first = False
        else:
            json_str += ","

        if value:
            json_str += '"%s":%s' % (name, json.dumps(value))

    json_str += "}"

    return json_str


# Visibility values for the UI
# Note:  could use enum in Python 3


class Visibility(object):
    DEFAULT = "DEFAULT"
    NON_SHAREABLE = "NON_SHARABLE"


class User(object):
    """
    Represents a user to TS.
    """

    def __init__(
        self,
        name,
        password=None,
        mail=None,
        display_name=None,
        group_names=None,
        visibility=Visibility.DEFAULT,
        created=None,
        id=None
    ):
        """
        Creates a new user object.
        :param name: The name of the user.  This is the login name.
        :type name: str
        :param password: The password for the user.  Not sent if there is no password.
        :type password: str
        :param mail: The user's email.
        :type mail: str
        :param display_name: The name to display in the UI.  Set to name if not provided.
        :type display_name: str
        :param group_names: Set of groups the user belongs to.
        :type group_names: list of str
        :param visibility: Visibility for sharing.
        :type visibility: str
        :param created: Epoch time when the user was created.
        :type created: str
        :param id: GUID from TS for the user.  Optional and not used for synching.
        :type created: str
        :return: Returns a new user object populated with the passed in values.
        :rtype: User
        """
        self.principalTypeEnum = "LOCAL_USER"
        self.name = name.strip()
        self.displayName = display_name if not None else name
        self.password = password
        self.mail = mail
        self.created = created
        if not group_names:
            self.groupNames = list()
        else:
            self.groupNames = list(group_names)
        self.visibility = visibility
        self.id = id

    def add_group(self, group_name):
        """
        Adds a parent group for the user.
        :param group_name: Name of the group to add.  Only one of any group is added.
        """

        if group_name not in self.groupNames:
            self.groupNames.append(group_name)

    def to_json(self):
        """
        Returns a JSON string representing this user.
        :return: A JSON string representing this user.
        :rtype: str
        """
        return obj_to_json(self)

    def __repr__(self):
        """
        Provides a string representation of the object.
        :return: A string representation of a user.
        :type: str
        """
        return self.to_json()


class Group(object):
    """
    Represents a group to TS.
    """

    def __init__(
        self,
        name,
        display_name=None,
        description=None,
        group_names=None,
        visibility=Visibility.DEFAULT,
        created=None,
    ):
        """
        Creates a new group object.
        :param name: The name of the group.
        :type name: str
        :param display_name: The name to display in the UI.  Set to name if not provided.
        :type display_name: str
        :param description: The optional description of the group.
        :type description: str
        :param group_names: Set of groups the group belongs to.
        :type group_names: list of str
        :param visibility: Indicates if the group is visibility or default.
        :type visibility: str
        :param created: The epoch date when the group was created.
        :type created: str
        :return: Returns a new group object populated with the passed in values.
        :rtype: Group
        """
        self.principalTypeEnum = "LOCAL_GROUP"
        self.name = name.strip()
        self.displayName = display_name if not None else name
        self.description = description
        self.visibility = visibility
        self.created = created
        if not group_names:
            self.groupNames = list()
        else:
            self.groupNames = list(group_names)

    def add_group(self, group_name):
        """
        Adds a parent group for the user.
        :param group_name: Name of the group to add.  Only one of any group is added.
        """

        if group_name not in self.groupNames:
            self.groupNames.append(group_name)

    def to_json(self):
        """
        Returns a JSON string representing this group.
        :return: A JSON string representing this group.
        :rtype: str
        """
        return obj_to_json(self)

    def __repr__(self):
        """
        Provides a string representation of the object.
        :return: A string representation of a group.
        :rtype: str
        """
        return self.to_json()


# Global list of users and groups. ----------------------------------------------------------------------

ValidationResults = namedtuple("ValidationResults", ["result", "issues"])


class UsersAndGroups(object):
    """
    Container for created users and groups.  Can be converted to JSON and sent to the TS API.
    """

    # Flags to say what to do when adding duplicates for users or groups.  Default is RAISE_ERROR_ON_DUPLICATE.
    RAISE_ERROR_ON_DUPLICATE = 0
    IGNORE_ON_DUPLICATE = 1
    OVERWRITE_ON_DUPLICATE = 2
    UPDATE_ON_DUPLICATE = 3

    def __init__(self):
        """
        Creates a new container for users and groups.
        """
        self.users = OrderedDict()
        self.groups = OrderedDict()

    def add_user(self, u, duplicate=RAISE_ERROR_ON_DUPLICATE):
        """
        Adds a user to the container.  Note that this does not make a copy of the user.
        :param u: User object to add to the container.
        :param duplicate: Flag to indicate how to handle duplicates.
        """
        user = self.get_user(u.name)
        if not user:
            self.users[u.name] = u
        else:
            if duplicate == UsersAndGroups.RAISE_ERROR_ON_DUPLICATE:
                raise Exception("Duplicate user %s" % u)

            elif duplicate == UsersAndGroups.IGNORE_ON_DUPLICATE:
                pass  # keep the old one.
            elif duplicate == UsersAndGroups.OVERWRITE_ON_DUPLICATE:
                self.users[u.name] = u
            elif duplicate == UsersAndGroups.UPDATE_ON_DUPLICATE:
                u.groupNames.extend(user.groupNames)
                self.users[u.name] = u
            else:
                raise Exception("Unkown duplication rule %s" % duplicate)

    def has_user(self, user_name):
        """
        Returns true if the user is in the collection.
        :param user_name: Name of the user to check for.
        :return: True if the user is in the set, else false.
        :rtype: bool
        """
        return user_name in self.users

    def get_user(self, user_name):
        """
        Returns the user with the given name or None.
        :param user_name: Name of the user to get.
        :return: The user object or None.
        :rtype: User
        """
        return self.users.get(user_name, None)

    def remove_user(self, user_name):
        """
        Removes the user with the given name if it is in the collection..
        :param user_name: Name of the user to remove.
        """
        return self.users.pop(user_name, None)

    def number_users(self):
        """
        Returns the number of users in the list.
        :return: The number of users in the list.
        :rtype: int
        """
        return len(self.users)

    def get_users(self):
        """
        Returns a list with all the users.  This is a copy so the users here won't be changed.
        :return:  The list of users.
        :rtype: list
        """
        return list(self.users.values())

    def add_group(self, g, duplicate=RAISE_ERROR_ON_DUPLICATE):
        """
        Adds a group to the container.  Note that this does not make a copy of the group.
        :param g: User object to add to the container.
        :param duplicate: Flag for what to do if there is a duplicate entry.
        """
        group = self.get_group(g.name)
        if not group:
            self.groups[g.name] = g
        else:
            if duplicate == UsersAndGroups.RAISE_ERROR_ON_DUPLICATE:
                raise Exception("Duplicate group %s" % g)

            elif duplicate == UsersAndGroups.OVERWRITE_ON_DUPLICATE:
                self.groups[g.name] = g
            elif duplicate == UsersAndGroups.UPDATE_ON_DUPLICATE:
                group.groupNames.extend(g.groupNames)
            else:
                raise Exception("Unkown duplication rule %s" % duplicate)

    def has_group(self, group_name):
        """
        Returns true if the group is in the collection.
        :param group_name: Name of the group to check for.
        :return: True if the group is in the set, else false.
        :rtype: bool
        """
        # print self.groups
        return group_name in self.groups

    def get_group(self, group_name):
        """
        Returns the group with the given name or None.
        param group_name: Name of the group to get.
        :return: The group object or None.
        :rtype: Group
        """
        return self.groups.get(group_name, None)

    def remove_group(self, group_name):
        """
        Removes the group with the given name if it is in the collection..
        :param group_name: Name of the group to remove.
        """
        return self.groups.pop(group_name, None)

    def number_groups(self):
        """
        Returns the number of groups in the list.
        :return: The number of groups in the list.
        :rtype: int
        """
        return len(self.groups)

    def get_groups(self):
        """
        Returns a list with all the groups.  This is a copy so the groups here won't be changed.
        :return:  The list of groups.
        :rtype: list
        """
        return list(self.groups.values())

    def to_json(self):
        """
        Adds a group to the container.
        :return: A JSON string representation the users and groups.
        :rtype: str
        """
        json_str = "["
        first = True
        for g in self.groups.values():
            if first:
                first = False
            else:
                json_str += ","
            json_str += g.to_json()

        for u in self.users.values():
            if first:
                first = False
            else:
                json_str += ","
            json_str += u.to_json()

        json_str += "]"

        return json_str

    def __repr__(self):
        """
        Retruns a string representation of the list.
        :return: A string representation of the list.
        :rtype: str
        """
        return self.to_json()

    def is_valid(self):
        """
        This method checks to see if the users and groups line up.  It makes sure that groups users belong to exist
        and that groups other groups belong to also exist.
        :return: Tuple with true or false if valid and list of errors if not.
        :rtype: (bool, issues)
        """
        issues = []
        valid = True  # true by default until an issue is found.

        for user in self.users.values():
            for parent_group in user.groupNames:
                if parent_group not in self.groups:
                    issue = "user group %s for user %s does not exist" % (
                        parent_group, user.name
                    )
                    print(issue)
                    issues.append(issue)
                    valid = False

        for group in self.groups.values():
            for parent_group in group.groupNames:
                if parent_group not in self.groups:
                    issue = "parent group %s for group %s does not exist" % (
                        parent_group, group.name
                    )
                    print(issue)
                    issues.append(issue)
                    valid = False

        return ValidationResults(result=valid, issues=issues)


class UGJsonReader(object):
    """
    Reads a user / group structure from JSON and returns a UserGroup object.
    """

    def read_from_file(self, filename):
        """
        Reads the JSON data from a file.
        :param filename: Name of the file to read.
        :type filename: str
        :return: A UsersAndGroups container based on the JSON.
        :rtype: UsersAndGroups
        """
        with open(filename, "r") as json_file:
            json_list = json.load(json_file)
            return self.parse_json(json_list)

    def read_from_string(self, json_string):
        """
        Reads the users and groups from a JSON string.
        :param json_string: String containing the JSON.
        :type json_string: str
        :return: A UsersAndGroups container based on the JSON.
        :rtype: UsersAndGroups
        """
        json_list = json.loads(json_string)
        return self.parse_json(json_list)

    def parse_json(self, json_list):
        """
        Parses a JSON list and creates a UserAndGroup object.
        :param json_list: List of JSON objects that represent users and groups.
        :returns: A user and group container with the users and groups.
        :rtype: UsersAndGroups
        """
        auag = UsersAndGroups()
        for value in json_list:
            if str(value["principalTypeEnum"]).endswith("_USER"):
                user = User(
                    name=value.get("name", None),
                    display_name=value.get("displayName", None),
                    mail=value.get("mail", None),
                    group_names=value.get("groupNames", None),
                    visibility=value.get("visibility", None),
                    created=value.get("created", None),
                    id=value.get("id", None)
                )
                auag.add_user(user)
            else:
                group = Group(
                    name=value.get("name", None),
                    display_name=value.get("displayName", None),
                    description=value.get("description", None),
                    group_names=value.get("groupNames", None),
                    visibility=value.get("visibility", None),
                )
                auag.add_group(group)
        return auag


def api_call(f):
    """
    Makes sure to try to call login if not already logged in.  This only works for classes that extend BaseApiInterface.
    :param f: Function to decorate.
    :return: A new callable method that will try to login first.
    """

    def wrap(self, *args, **kwargs):
        """
        Verifies that the user is logged in and then makes the call.  Assumes something will be returned.
        :param self:  Instance calling a method.
        :param args:  Place arguments.
        :param kwargs: Key word arguments.
        :return: Whatever the wrapped method returns.
        """
        if not self.is_authenticated():
            self.login()
        return f(self, *args, **kwargs)

    return wrap


class BaseApiInterface(object):
    """
    Provides basic support for calling the ThoughtSpot APIs, particularly for logging in.
    """
    SERVER_URL = "{tsurl}/callosum/v1"

    def __init__(self, tsurl, username, password, disable_ssl=False):
        """
        Creates a new sync object and logs into ThoughtSpot
        :param tsurl: Root ThoughtSpot URL, e.g. http://some-company.com/
        :type tsurl: str
        :param username: Name of the admin login to use.
        :type username: str
        :param password: Password for admin login.
        :type password: str
        :param disable_ssl: If true, then disable SSL for calls.
        password for all users.  This can be significantly faster than individual passwords.
        """
        self.tsurl = tsurl
        self.username = username
        self.password = password
        self.cookies = None
        self.session = requests.Session()
        if disable_ssl:
            requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
            self.session.verify = False
        self.session.headers = {"X-Requested-By": "ThoughtSpot"}

    def login(self):
        """
        Log into the ThoughtSpot server.
        """
        url = self.format_url(SyncUserAndGroups.LOGIN_URL)
        response = self.session.post(
            url, data={"username": self.username, "password": self.password}
        )

        if response.status_code == 204:
            self.cookies = response.cookies
            logging.info("Successfully logged in as %s" % self.username)
        else:
            logging.error("Failed to log in as %s" % self.username)
            raise requests.ConnectionError(
                "Error logging in to TS (%d)" % response.status_code,
                response.text,
            )

    def is_authenticated(self):
        """
        Returns true if the session is authenticated
        :return: True if the session is authenticated.
        :rtype: bool
        """
        return self.cookies is not None

    def format_url(self, url):
        """
        Returns a URL that has the correct server.
        :param url: The URL template to add the server to.
        :type url: str
        :return: A URL that has the correct server info.
        :rtype: str
        """
        url = BaseApiInterface.SERVER_URL + url
        return url.format(tsurl=self.tsurl)


class SyncUserAndGroups(BaseApiInterface):
    """
    Synchronized with ThoughtSpot and also gets users and groups from ThoughtSpot.
    """

    LOGIN_URL = "/tspublic/v1/session/login"
    GET_ALL_URL = "/tspublic/v1/user/list"
    SYNC_ALL_URL = "/tspublic/v1/user/sync"
    UPDATE_PASSWORD_URL = "/tspublic/v1/user/updatepassword"
    DELETE_USERS_URL = "/session/user/deleteusers"
    DELETE_GROUPS_URL = "/session/group/deletegroups"
    USER_METADATA_URL = "/tspublic/v1/metadata/listobjectheaders?type=USER"
    GROUP_METADATA_URL = "/tspublic/v1/metadata/listobjectheaders?type=USER_GROUP"

    def __init__(
        self,
        tsurl,
        username,
        password,
        disable_ssl=False,
        global_password=False,
    ):
        """
        Creates a new sync object and logs into ThoughtSpot
        :param tsurl: Root ThoughtSpot URL, e.g. http://some-company.com/
        :param username: Name of the admin login to use.
        :param password: Password for admin login.
        :param disable_ssl: If true, then disable SSL for calls.
        :param global_password: If provided, will be passed to the sync call.  This is used to have a single
        password for all users.  This can be significantly faster than individual passwords.
        """
        super(SyncUserAndGroups, self).__init__(
            tsurl=tsurl,
            username=username,
            password=password,
            disable_ssl=disable_ssl,
        )
        self.global_password = global_password

    @api_call
    def get_all_users_and_groups(self):
        """
        Returns all users and groups from the server.
        :return: All users and groups from the server.
        :rtype: UsersAndGroups
        """

        url = self.format_url(SyncUserAndGroups.GET_ALL_URL)
        response = self.session.get(url, cookies=self.cookies)
        if response.status_code == 200:
            logging.info("Successfully got users and groups.")
            json_list = json.loads(response.text)
            reader = UGJsonReader()
            auag = reader.parse_json(json_list=json_list)
            return auag

        else:
            logging.error("Failed to get users and groups.")
            raise requests.ConnectionError(
                "Error getting users and groups (%d)" % response.status_code,
                response.text,
            )

    @api_call
    def get_user_metadata(self):
        """
        Returns a list of User objects based on the metadata.
        :return: A list of user objects.
        :rtype: list of User
        """
        url = self.format_url(SyncUserAndGroups.USER_METADATA_URL)
        response = self.session.get(url, cookies=self.cookies)
        users = []
        if response.status_code == 200:
            logging.info("Successfully got user metadata.")
            json_list = json.loads(response.text)
            for value in json_list:
                user = User(
                    name=value.get("name", None),
                    display_name=value.get("displayName", None),
                    mail=value.get("mail", None),
                    group_names=value.get("groupNames", None),
                    visibility=value.get("visibility", None),
                    created=value.get("created", None),
                    id=value.get("id", None)
                )
                users.append(user)
            return users

        else:
            logging.error("Failed to get user metadata.")
            raise requests.ConnectionError(
                "Error getting user metadata (%d)" % response.status_code,
                response.text,
                )

    @api_call
    def sync_users_and_groups(
        self, users_and_groups, apply_changes=True, remove_deleted=False
    ):
        """
        Syncs users and groups.
        :param users_and_groups: List of users and groups to sync.
        :type users_and_groups: UsersAndGroups
        :param apply_changes: If true, changes will be applied.  If not, then it just says what will happen.
        :type apply_changes: bool
        :param remove_deleted: Flag to removed deleted users.  If true, delete.
        :type remove_deleted: bool
        :returns: The response from the sync.
        """

        is_valid = users_and_groups.is_valid()
        if not is_valid[0]:
            # print("Invalid user and group structure.")
            raise Exception("Invalid users and groups")

        url = self.format_url(SyncUserAndGroups.SYNC_ALL_URL)

        logging.debug("calling %s" % url)
        json_str = users_and_groups.to_json()
        logging.debug("data == %s" % json_str)
        json.loads(json_str)  # do a load to see if it breaks due to bad JSON.

        tmp_file = "/tmp/ug.json.%d" % time.time()
        with open(tmp_file, "w") as out:
            out.write(json_str)

        params = {
            "principals": (tmp_file, open(tmp_file, "rb"), "text/json"),
            "applyChanges": json.dumps(apply_changes),
            "removeDeleted": json.dumps(remove_deleted),
        }

        if self.global_password:
            params["password"] = json.dumps(self.global_password)

        response = self.session.post(url, files=params, cookies=self.cookies)

        if response.status_code == 200:
            logging.info("Successfully synced users and groups.")
            print(response.text.encode("utf-8"))
            return response

        else:
            logging.error("Failed synced users and groups.")
            print(response.text.encode("utf-8"))
            with open("ts_users_and_groups.json", "w") as outfile:
                outfile.write(json_str.encode("utf-8"))
            raise requests.ConnectionError(
                "Error syncing users and groups (%d)" % response.status_code,
                response.text,
            )

    @api_call
    def delete_users(self, usernames):
        """
        Deletes a list of users based on their user name.
        :param usernames: List of the names of the users to delete.
        :type usernames: list of str
        """

        # for each username, get the guid and put in a list.  Log errors for users not found, but don't stop.
        url = self.format_url(SyncUserAndGroups.USER_METADATA_URL)
        response = self.session.get(url, cookies=self.cookies)
        users = {}
        if response.status_code == 200:
            logging.info("Successfully got user metadata.")
            json_list = json.loads(response.text)
            for h in json_list:
                name = h["name"]
                id = h["id"]
                users[name] = id

            user_list = []
            for u in usernames:
                id = users.get(u, None)
                if not id:
                    eprint(
                        "WARNING:  user %s not found, not attempting to delete this user."
                        % u
                    )
                else:
                    user_list.append(id)

            if not user_list:
                eprint("No valid users to delete.")
                return

            url = self.format_url(SyncUserAndGroups.DELETE_USERS_URL)
            params = {"ids": json.dumps(user_list)}
            response = self.session.post(
                url, data=params, cookies=self.cookies
            )

            if response.status_code != 204:
                logging.error("Failed to delete %s" % user_list)
                raise requests.ConnectionError(
                    "Error getting users and groups (%d)"
                    % response.status_code,
                    response.text,
                )

        else:
            logging.error("Failed to get users and groups.")
            raise requests.ConnectionError(
                "Error getting users and groups (%d)" % response.status_code,
                response.text,
            )

    def delete_user(self, username):
        """
        Deletes the user with the given username.
        :param username: The name of the user.
        :type username: str
        """
        self.delete_users([username])  # just call the list method.

    @api_call
    def delete_groups(self, groupnames):
        """
        Deletes a list of groups based on their group name.
        :param groupnames: List of the names of the groups to delete.
        :type groupnames: list of str
        """

        # for each groupname, get the guid and put in a list.  Log errors for groups not found, but don't stop.
        url = self.format_url(SyncUserAndGroups.GROUP_METADATA_URL)
        response = self.session.get(url, cookies=self.cookies)
        groups = {}
        if response.status_code == 200:
            logging.info("Successfully got group metadata.")
            json_list = json.loads(response.text)
            # for h in json_list["headers"]:
            for h in json_list:
                name = h["name"]
                id = h["id"]
                groups[name] = id

            group_list = []
            for u in groupnames:
                id = groups.get(u, None)
                if not id:
                    eprint(
                        "WARNING:  group %s not found, not attempting to delete this group."
                        % u
                    )
                else:
                    group_list.append(id)

            if not group_list:
                eprint("No valid groups to delete.")
                return

            url = self.format_url(SyncUserAndGroups.DELETE_GROUPS_URL)
            params = {"ids": json.dumps(group_list)}
            response = self.session.post(
                url, data=params, cookies=self.cookies
            )

            if response.status_code != 204:
                logging.error("Failed to delete %s" % group_list)
                raise requests.ConnectionError(
                    "Error getting groups and groups (%d)"
                    % response.status_code,
                    response.text,
                )

        else:
            logging.error("Failed to get users and groups.")
            raise requests.ConnectionError(
                "Error getting users and groups (%d)" % response.status_code,
                response.text,
            )

    def delete_group(self, groupname):
        """
        Deletes the group with the given groupname.
        :param groupname: The name of the group.
        :type groupname: str
        """
        self.delete_groups([groupname])  # just call the list method.

    @api_call
    def update_user_password(self, userid, currentpassword, password):
        """
        Updates the password for a user.
        :param userid: User id for the user to change the password for.
        :type userid: str
        :param currentpassword: Previous password for the user.
        :type currentpassword: str
        :param password: New password for the user.
        :type password: str
        """

        url = self.format_url(SyncUserAndGroups.UPDATE_PASSWORD_URL)
        params = {
            "userid": userid,
            "currentpassword": {
                "password": [currentpassword], "empty": "false"
            },
            "password": {"password": [password], "empty": "false"},
        }

        params = json.dumps(params)
        print(params)

        return  # TODO add after 4.4 is released.

        response = self.session.post(url, data=params, cookies=self.cookies)

        if response.status_code == 200:
            logging.info("Successfully updated password for %s." % userid)
        else:
            logging.error("Failed to update password for %s." % userid)
            raise requests.ConnectionError(
                "Error (%d) updating user password for %s:  %s"
                % (response.status_code, userid, response.text)
            )


class Privileges(object):
    """
    Contains the various privileges that groups can have.
    """
    IS_ADMINSTRATOR = "ADMINISTRATION"
    CAN_UPLOAD_DATA = "USERDATAUPLOADING"
    CAN_DOWNLOAD_DATA = "DATADOWNLOADING"
    CAN_SHARE_WITH_ALL = "SHAREWITHALL"
    CAN_MANAGE_DATA = "DATAMANAGEMENT"
    CAN_SCHEDULE_PINBOARDS = "JOBSCHEDULING"
    CAN_USE_SPOTIQ = "A3ANALYSIS"
    CAN_ADMINISTER_RLS = "BYPASSRLS"
    CAN_AUTHOR = "AUTHORING"
    CAN_MANAGE_SYSTEM = "SYSTEMMANAGEMENT"


class SetGroupPrivilegesAPI(BaseApiInterface):

    # Note that some of these URLs are not part of the public API and subject to change.
    METADATA_LIST_URL = "/tspublic/v1/metadata/listobjectheaders?type=USER_GROUP"
    METADATA_DETAIL_URL = "/metadata/detail/{guid}?type=USER_GROUP"

    ADD_PRIVILEGE_URL = "/tspublic/v1/group/addprivilege"
    REMOVE_PRIVILEGE_URL = "/tspublic/v1/group/removeprivilege"

    def __init__(self, tsurl, username, password, disable_ssl=False):
        """
        Creates a new sync object and logs into ThoughtSpot
        :param tsurl: Root ThoughtSpot URL, e.g. http://some-company.com/
        :param username: Name of the admin login to use.
        :param password: Password for admin login.
        :param disable_ssl: If true, then disable SSL for calls.
        """
        super(SetGroupPrivilegesAPI, self).__init__(
            tsurl=tsurl,
            username=username,
            password=password,
            disable_ssl=disable_ssl,
        )

    @api_call
    def get_privileges_for_group(self, group_name):
        """
        Gets the current privileges for a given group.
        :param group_name:  Name of the group to get privileges for.
        :returns: A list of privileges.
        :rtype: list of str
        """
        url = self.format_url(
            SetGroupPrivilegesAPI.METADATA_LIST_URL
        ) + "&pattern=" + group_name
        response = self.session.get(url, cookies=self.cookies)
        if response.status_code == 200:  # success
            results = json.loads(response.text)
            try:
                group_id = results[0][
                    "id"
                ]  # should always be present, but might want to add try / catch.
                detail_url = SetGroupPrivilegesAPI.METADATA_DETAIL_URL.format(
                    guid=group_id
                )
                detail_url = self.format_url(detail_url)
                detail_response = self.session.get(
                    detail_url, cookies=self.cookies
                )
                if detail_response.status_code == 200:  # success
                    privileges = json.loads(detail_response.text)["privileges"]
                    return privileges

                else:
                    logging.error(
                        "Failed to get privileges for group %s" % group_name
                    )
                    raise requests.ConnectionError(
                        "Error (%d) setting privileges for group %s.  %s"
                        % (response.status_code, group_name, response.text)
                    )

            except Exception:
                print("Error getting group details.")
                raise

        else:
            logging.error("Failed to get privileges for group %s" % group_name)
            raise requests.ConnectionError(
                "Error (%d) setting privileges for group %s.  %s"
                % (response.status_code, group_name, response.text)
            )

    @api_call
    def add_privilege(self, groups, privilege):
        """
        Adds a privilege to a list of groups.
        :param groups List of groups to add the privilege to.
        :type groups: list of str
        :param privilege: Privilege being set.
        :type privilege: str
        """

        url = self.format_url(SetGroupPrivilegesAPI.ADD_PRIVILEGE_URL)

        params = {"privilege": privilege, "groupNames": json.dumps(groups)}
        response = self.session.post(url, files=params, cookies=self.cookies)

        if response.status_code == 204:
            logging.info(
                "Successfully added privilege %s for groups %s."
                % (privilege, groups)
            )
        else:
            logging.error(
                "Failed to add privilege %s for groups %s."
                % (privilege, groups)
            )
            raise requests.ConnectionError(
                "Error (%d) adding privilege %s for groups %s.  %s"
                % (response.status_code, privilege, groups, response.text)
            )

    @api_call
    def remove_privilege(self, groups, privilege):
        """
        Removes a privilege to a list of groups.
        :param groups List of groups to add the privilege to.
        :type groups: list of str
        :param privilege: Privilege being removed.
        :type privilege: str
        """

        url = self.format_url(SetGroupPrivilegesAPI.REMOVE_PRIVILEGE_URL)

        params = {"privilege": privilege, "groupNames": json.dumps(groups)}
        response = self.session.post(url, files=params, cookies=self.cookies)

        if response.status_code == 204:
            logging.info(
                "Successfully removed privilege %s for groups %s."
                % (privilege, groups)
            )
        else:
            logging.error(
                "Failed to remove privilege %s for groups %s."
                % (privilege, groups)
            )
            raise requests.ConnectionError(
                "Error (%d) removing privilege %s for groups %s.  %s"
                % (response.status_code, privilege, groups, response.text)
            )


class TransferOwnershipApi(BaseApiInterface):

    TRANSFER_OWNERSHIP_URL = "{tsurl}/tspublic/v1/user/transfer/ownership"

    def __init__(self, tsurl, username, password, disable_ssl=False):
        """
        Creates a new sync object and logs into ThoughtSpot
        :param tsurl: Root ThoughtSpot URL, e.g. http://some-company.com/
        :param username: Name of the admin login to use.
        :param password: Password for admin login.
        :param disable_ssl: If true, then disable SSL for calls.
        """
        super(TransferOwnershipApi, self).__init__(
            tsurl=tsurl,
            username=username,
            password=password,
            disable_ssl=disable_ssl,
        )

    @api_call
    def transfer_ownership(self, from_username, to_username):
        """
        Transfer ownersip of all objects from one user to another.
        :param from_username: User name for the user to change the ownership for.
        :type from_username: str
        :param to_username: User name for the user to change the ownership to.
        :type to_username: str
        """

        url = self.format_url(TransferOwnershipApi.TRANSFER_OWNERSHIP_URL)
        url = url + "?fromUserName=" + from_username + "&toUserName=" + to_username
        response = self.session.post(url, cookies=self.cookies)

        if response.status_code == 204:
            logging.info(
                "Successfully transferred ownership to %s." % to_username
            )
        else:
            logging.error("Failed to transfer ownership to %s." % to_username)
            raise requests.ConnectionError(
                "Error (%d) transferring  ownership to %s:  %s"
                % (response.status_code, to_username, response.text)
            )


class UGXLSWriter(object):
    """
    Writes users and groups to an Excel spreadsheet.
    """

    def write(self, users_and_groups, filename):
        """
        Writes the content to the given file.
        :param users_and_groups:  The UsersAndGroups object to write.
        :type users_and_groups: UsersAndGroups
        :param filename:  Name of the file to write to.  No extension is expected and one will be added.
        :type filename: str
        """
        workbook = Workbook()
        workbook.remove_sheet(
            workbook.active
        )  # remove the default sheet since we'll be creating the ones we want.
        self._write_users(workbook, users_and_groups.get_users())
        self._write_groups(workbook, users_and_groups.get_groups())
        if not (filename.endswith("xls") or filename.endswith("xlsx")):
            filename += ".xlsx"

        workbook.save(filename)

    def _write_users(self, workbook, users):
        """
        Writes the users to a worksheet.
        :param workbook:  The workbook to write to.
        :type workbook:  Workbook
        :param users:  The list of groups to write.
        :type users: list of User
        :return:
        """
        ws = workbook.create_sheet(title="Users")
        self._write_header(
            ws,
            [
                "Name",
                "Password",
                "Display Name",
                "Email",
                "Groups",
                "Visibility",
                "Created",
                "ID"
            ],
        )
        cnt = 2  # start after header.
        for user in users:
            ws.cell(column=1, row=cnt, value=user.name)
            ws.cell(column=2, row=cnt, value=user.password)
            ws.cell(column=3, row=cnt, value=user.displayName)
            ws.cell(column=4, row=cnt, value=user.mail)
            ws.cell(column=5, row=cnt, value=json.dumps(user.groupNames))
            ws.cell(column=6, row=cnt, value=user.visibility)
            ws.cell(column=7, row=cnt, value=user.created)
            ws.cell(column=8, row=cnt, value=user.id)
            cnt += 1

    def _write_groups(self, workbook, groups):
        """
        Writes the groups to a worksheet.
        :param workbook:  The workbook to write to.
        :type workbook:  Workbook
        :param groups:  The list of groups to write.
        :type groups: list
        :return:
        """
        ws = workbook.create_sheet(title="Groups")
        self._write_header(
            ws,
            [
                "Name",
                "Display Name",
                "Description",
                "Groups",
                "Visibility",
                "Created",
            ],
        )
        cnt = 2  # start after header.
        for group in groups:
            ws.cell(column=1, row=cnt, value=group.name)
            ws.cell(column=2, row=cnt, value=group.displayName)
            ws.cell(column=3, row=cnt, value=group.description)
            ws.cell(column=4, row=cnt, value=json.dumps(group.groupNames))
            ws.cell(column=5, row=cnt, value=group.visibility)
            ws.cell(column=6, row=cnt, value=group.created)
            cnt += 1

    def _write_header(self, worksheet, cols):
        """
        Writes the header for the given worksheet in row 1.
        :param worksheet:  Worksheet to write to.
        :param cols:  List of columns to write.
        """
        for ccnt in range(0, len(cols)):
            worksheet.cell(column=(ccnt + 1), row=1, value=cols[ccnt])


class UGXLSReader(object):
    """
    Reads user and group info from an Excel file that is formatted the same as the UGXLSWriter writes.
    """

    required_sheets = ["Users", "Groups"]
    required_columns = {
        "Users": [
            "Name",
            "Password",
            "Display Name",
            "Email",
            "Groups",
            "Visibility"
        ],
        "Groups": [
            "Name",
            "Display Name",
            "Description",
            "Groups",
            "Visibility"
        ],
    }

    def __init__(self):
        """
        Creates a new UGXLSReader
        """
        self.workbook = None
        self.indices = {}
        self.users_and_groups = UsersAndGroups()

    def read_from_excel(self, filepath):
        """
        Reads users and groups from the given file.
        :param filepath:  Path to the Excel file to read from.
        :type filepath: str
        :return: Returns the users and groups read from the Excel file.  The users and groups are not validated
        :rtype UsersAndGroups
        so that they can be modified prior to validation.
        """
        self.workbook = xlrd.open_workbook(filepath)
        if self._verify_file_format():
            self._get_column_indices()
            self._read_users_from_workbook()
            self._read_groups_from_workbook()
        return self.users_and_groups

    def _verify_file_format(self):
        """
        :return: True if the format of the workbook is valid.
        :rtype: bool
        """
        is_valid = True
        sheet_names = self.workbook.sheet_names()
        for required_sheet in UGXLSReader.required_sheets:
            if required_sheet not in sheet_names:
                eprint("Error:  missing sheet %s!" % required_sheet)
                is_valid = False
            else:
                sheet = self.workbook.sheet_by_name(required_sheet)
                header_row = sheet.row_values(rowx=0, start_colx=0)
                for required_column in UGXLSReader.required_columns[
                    required_sheet
                ]:
                    if required_column not in header_row:
                        eprint(
                            "Error:  missing column %s in sheet %s!"
                            % (required_column, required_sheet)
                        )
                        is_valid = False

        return is_valid

    def _get_column_indices(self):
        """
        Reads the sheets to get all of the column indices.  Assumes the format was already checked.
        """
        sheet_names = self.workbook.sheet_names()
        for sheet_name in sheet_names:
            if sheet_name in self.required_sheets:
                sheet = self.workbook.sheet_by_name(sheet_name)
                col_indices = {}
                ccnt = 0
                for col in sheet.row_values(rowx=0, start_colx=0):
                    col_indices[col] = ccnt
                    ccnt += 1
                self.indices[sheet_name] = col_indices

    def _read_users_from_workbook(self):
        """
        Reads all the users from the workbook.
        """

        table_sheet = self.workbook.sheet_by_name("Users")
        indices = self.indices["Users"]

        for row_count in range(1, table_sheet.nrows):
            row = table_sheet.row_values(rowx=row_count, start_colx=0)

            # "Name", "Password", "Display Name", "Email", "Description", "Groups", "Visibility"
            username = row[indices["Name"]]
            password = row[indices["Password"]]
            display_name = row[indices["Display Name"]]
            email = row[indices["Email"]]
            groups = []
            if row[indices["Groups"]] and row[
                indices["Groups"]
            ]:
                groups = ast.literal_eval(
                    row[indices["Groups"]]
                )  # assumes a valid list format, e.g. ["a", "b", ...]
            visibility = row[indices["Visibility"]]

            try:
                user = User(
                    name=username,
                    password=password,
                    display_name=display_name,
                    mail=email,
                    group_names=groups,
                    visibility=visibility,
                )
                # The format should be consistent with only one user per line.
                self.users_and_groups.add_user(
                    user, duplicate=UsersAndGroups.RAISE_ERROR_ON_DUPLICATE
                )
            except:
                eprint("Error reading user with name %s" % username)

    def _read_groups_from_workbook(self):
        """
        Reads all the groups from the workbook.
        """

        table_sheet = self.workbook.sheet_by_name("Groups")
        indices = self.indices["Groups"]

        for row_count in range(1, table_sheet.nrows):
            row = table_sheet.row_values(rowx=row_count, start_colx=0)

            # Name", "Display Name", "Description", "Groups", "Visibility"
            group_name = row[indices["Name"]]
            display_name = row[indices["Display Name"]]
            description = row[indices["Description"]]
            visibility = row[indices["Visibility"]]

            groups = []
            if row[indices["Groups"]] and row[
                indices["Groups"]
            ]:
                groups = ast.literal_eval(
                    row[indices["Groups"]]
                )  # assumes a valid list format, e.g. ["a", "b", ...]
            try:
                group = Group(
                    name=group_name,
                    display_name=display_name,
                    description=description,
                    group_names=groups,
                    visibility=visibility,
                )
                # The format should be consistent with only one group per line.
                self.users_and_groups.add_group(
                    group, duplicate=UsersAndGroups.RAISE_ERROR_ON_DUPLICATE
                )
            except:
                eprint("Error reading group with name %s" % group_name)
